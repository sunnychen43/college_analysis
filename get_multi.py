from openpyxl import load_workbook
import subprocess as sp
import sys
import db

# get_multi.py refrence_file_name new_file_name
if __name__ == '__main__':
	if(len(sys.argv) != 3):
		print('Invalid arguments')
		quit()
	try:
		refrence_wb = load_workbook(sys.argv[1])
	except FileNotFoundError:
		print('Refrence File not Found')
		quit()

	refrence_ws = refrence_wb.active
	max_row = refrence_ws.max_row - 1
	tmp = sp.call('clear', shell=True)  # Clear screen

	for i in range(max_row):
		row = i + 2
		school = refrence_ws['A' + str(row)].value  # Get information for each thread
		year = refrence_ws['B' + str(row)].value
		print(school, year, i + 1, '/', max_row)
		if refrence_ws['C' + str(row)].value is not None:  # If there is a link
			try:
				db.create(refrence_ws['C' + str(row)].value, sys.argv[2], school, year, 'ED')
			except IndexError:  # Don't remember why this is here, but it accounts for some error
				continue
		if refrence_ws['D' + str(row)].value is not None:
			try:
				db.create(refrence_ws['D' + str(row)].value, sys.argv[2], school, year, 'RD')
			except IndexError:
				continue
		tmp = sp.call('clear', shell=True)
	wb = load_workbook(sys.argv[2])  # Run formatting on completed workbook
	wb = db.clean(wb)
	db.parse_sat1(wb)
	db.parse_act(wb)
	db.create_additional_entries(wb)
	wb.save(sys.argv[2])
