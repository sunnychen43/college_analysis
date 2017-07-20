from lxml import html
from openpyxl import workbook
from openpyxl import load_workbook
import requests
import re
import comment
import classifier_lists


# Imports comment_set into workbook
def save_comment_set(comment_set, wb, id, row):
    ws = wb.active
    comment_count = 1  # Counter

    for single_comment in comment_set:  # Loops through all comments in the set
        single_comment = comment.format(single_comment)

        if not single_comment:  # Skip invalid comments
            comment_count += 1
            continue

        for data_point in single_comment:  # Imports data in comment to workbook
            id_cell = ws["A" + str(row)]
            id_cell.value = id

            data_cell = ws["C" + str(row)]
            data_cell.value = data_point
            row += 1

        id += 1
        row += 1
        comment_count += 1
    return id, row  # Return the next blank lines


# Save data from page into workbook
def get_comments(url, wb):
    page = requests.get(url)
    tree = html.fromstring(page.content)

    nav_container = tree.xpath('//*[@id="PagerBefore"]/node()')  # Locates link of last page in thread
    last_page = nav_container[len(nav_container) - 4]
    last_page_url = last_page.attrib['href']

    p = re.compile('(?:-p)(\d+)')  # Extracts the last page number (-p12345)
    m = re.findall(p, last_page_url)
    num_pages = int(m[0])

    wb.active['A1'].value = "ID:"
    wb.active['B1'].value = "Class:"
    wb.active['C1'].value = "Data:"

    id = 1
    row = 2
    for page_id in range(num_pages):  # Saves comments for each page
        page_id += 1
        if page_id == 1:
            new_url = url
        else:
            new_url = url[:-5] + "-p" + str(page_id) + ".html"  # Modifies base url to include page number

        comment_set = comment.scrape_all_comments(new_url)
        id, row = save_comment_set(comment_set, wb, id, row)
        print('Save:', "Page", str(page_id) + '/' + str(num_pages), "done")


# Assigns each entry a class if it matches
def classify(wb):
    ws = wb.active
    local_regex_dict = classifier_lists.regex_dict
    max_row = ws.max_row

    for row in range(max_row):
        row += 1
        if row == 1:  # Skip headers
            continue

        data_cell = ws['C' + str(row)]
        data = data_cell.value
        if data is None:

            continue
        for key, value in local_regex_dict.items():
            p = re.compile(value)
            m = p.search(data)

            if m:
                ws['B' + str(row)].value = key
                data_cell.value = re.sub(p, '', data)  # Removes identifier from every cell in a category
                break
    print('Classify workbook done')


# Shrinks bulleted entries into one cell
def collate(wb):
    ws = wb.active
    max_row = ws.max_row
    in_list = False
    list_class = None
    list_data_cell = None
    str_list = []

    for row in range(max_row):
        row += 1
        row_class = ws['B' + str(row)].value
        row_data = ws['C' + str(row)].value

        if in_list:
            # Detects end of the list
            if row_class is not None and row_class != list_class:
                for string in str_list:
                    if string is None:
                        continue
                    try:
                        list_data_cell.value += ' ' + string
                    except TypeError:  # Original data cell doesn't exist
                        list_data_cell.value = string
                str_list = []
                in_list = False

            elif row_class is None:  # Add list element
                str_list.append(row_data)
                ws['C' + str(row)].value = None

        if not in_list:  # Check for start of list
            if row_class in classifier_lists.list_categories:
                list_class = row_class
                list_data_cell = ws['C' + str(row)]
                in_list = True
    print('Collate workbook done')