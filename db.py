from openpyxl import load_workbook
from openpyxl import workbook
import expand_entry
import classifier_lists
import excel
import re


# Create headers for database
def create_headers(wb):
	ws = wb.active
	counter = 1
	for header in classifier_lists.headers:
		ws.cell(column=counter, row=1).value = header
		counter += 1
	print('Create header done')


# Moves raw data into db
def raw_to_db(raw_wb, db_wb, college, year, type):
	raw_ws = raw_wb.active
	db_ws = db_wb.active
	max_db_row = db_ws.max_row
	num_entries = raw_ws.max_row - 1

	for entry in range(num_entries):
		raw_row = entry + 2  # Start scanning on the second line

		entry_id = raw_ws.cell(column=classifier_lists.headers.index('ID:') + 1, row=raw_row).value
		if entry_id is None:  # Blank line
			continue

		db_row = max_db_row + int(entry_id)  # Start at first free line

		if db_ws.cell(column=classifier_lists.headers.index('ID:') + 1, row=db_row).value is None:  # Create entry if it doesn't exist
			db_ws.cell(column=classifier_lists.headers.index('ID:') + 1, row=db_row).value = db_row - 1
			db_ws.cell(column=classifier_lists.headers.index('College:') + 1, row=db_row).value = college
			db_ws.cell(column=classifier_lists.headers.index('Year:') + 1, row=db_row).value = year
			db_ws.cell(column=classifier_lists.headers.index('Type:') + 1, row=db_row).value = type

		entry_class = raw_ws.cell(column=2, row=raw_row).value
		if entry_class is not None:
			try:
				db_column = classifier_lists.headers.index(entry_class) + 1  # Sets column under matching header
			except:  # If class is not appropriate header
				continue
			db_ws.cell(column=db_column, row=db_row).value = raw_ws.cell(column=3, row=raw_row).value  # Copies raw data to db
	print('Convert to db done')


# Main method, creates db
def create(url, file_name, school, year, type):
	raw_wb = workbook.Workbook()
	excel.get_comments(url, raw_wb)
	excel.classify(raw_wb)
	excel.collate(raw_wb)

	try:
		db_wb = load_workbook(file_name)
	except FileNotFoundError:
		db_wb = workbook.Workbook()
	create_headers(db_wb)
	raw_to_db(raw_wb, db_wb, school, year, type)
	db_wb.save(file_name)


# Removes blank lines
def clean(wb):
	old_wb = wb
	old_ws = old_wb.active
	new_wb = workbook.Workbook()
	new_ws = new_wb.active
	max_row = old_ws.max_row
	max_column = old_ws.max_column

	for i in range(max_row):
		row = i + 1
		if old_ws.cell(column=classifier_lists.headers.index('SAT I:') + 1, row=row).value is None:
			if old_ws.cell(column=classifier_lists.headers.index('SAT II:') + 1, row=row).value is None:
				if old_ws.cell(column=classifier_lists.headers.index('GPA:') + 1, row=row).value is None:
					continue
		for i in range(max_column):
			column = i + 1
			new_ws.cell(row=row, column=column).value = old_ws.cell(row=row, column=column).value
	print('Clean done')
	return new_wb


def parse_sat1(wb):
	ws = wb.active
	max_row = ws.max_row
	for i in range(max_row - 1):
		row = i + 2
		text = ws.cell(column=classifier_lists.headers.index('SAT I:') + 1, row=row).value
		if text is None:
			continue
		totalscorep = re.compile('(\d{3}0)')
		total_score = totalscorep.findall(text)

		subscorep = re.compile('((?<!\d)\d{2}0(?!\d))')
		subscores = subscorep.findall(text)
		subcatp = re.compile('(?=(math|reading|writing|(?<![a-z])w(?![a-z])|(?<![a-z])r(?![a-z])|(?<![a-z])cr(?![a-z])|(?<![a-z])m(?![a-z])))', re.IGNORECASE)
		subcats = subcatp.findall(text)

		breakdown = []
		if len(subcats) == 3 and len(subscores) == 3:
			if len(total_score) == 0:
				total_score.append(str(sum([int(score) for score in subscores])))
			for i in range(3):
				subject = subcats[i]
				if subject.lower() in ['cr', 'r', 'reading', 'critical reading']:
					breakdown.append(('Reading:', subscores[i]))
					continue
				if subject.lower() in ['m', 'math', 'mathematics']:
					breakdown.append(('Math:', subscores[i]))
					continue
				if subject.lower() in ['w', 'writing']:
					breakdown.append(('Writing:', subscores[i]))
					continue
		if len(total_score) != 0:
			ws.cell(column=classifier_lists.headers.index('Total Score') + 1, row=row).value = total_score[0]
		for subject in breakdown:
			ws.cell(column=classifier_lists.headers.index(subject[0]) + 1, row=row).value = subject[1]
	print('SAT I done')


def parse_act(wb):
	ws = wb.active
	max_row = ws.max_row
	for i in range(max_row - 1):
		row = i + 2
		text = ws.cell(column=classifier_lists.headers.index('ACT:') + 1, row=row).value
		if text is None:
			continue
		p = re.compile('(\d{2})')
		matches = p.findall(text)
		if len(matches) == 0:
			continue
		act_total = 0
		for match in matches:
			act_total += int(match)
		act_average = int(act_total / len(matches))
		ws.cell(column=classifier_lists.headers.index('ACT Composite:') + 1, row=row).value = str(act_average)
	print('ACT done')


def create_additional_entries(wb):
	ws = wb.active
	max_row = ws.max_row
	max_column = ws.max_column
	counter = 1  # Counter to find next blank ine
	for i in range(max_row - 1):  # Loop through all but first row
		row = i + 2
		decisions = expand_entry.get_decisions(ws.cell(column=classifier_lists.headers.index('Where else?') + 1, row=row).value)
		ws.cell(column=classifier_lists.headers.index('Where else?') + 1, row=row).value = None
		additional_schools = []
		if decisions is None:
			continue
		for decision, schools in decisions:
			for school in schools.split(', '):
				additional_schools.append((decision, school))
		if additional_schools is None:
			continue
		for decision, school in additional_schools:  # Add line per additional decision
			if school == '':
				continue
			new_row = max_row + counter
			for j in range(max_column - 1):  # Copy all columns but the last 4
				new_column = j + 1
				ws.cell(column=new_column, row=new_row).value = ws.cell(column=new_column, row=row).value  # Copy line to bottom of worksheet
			ws.cell(column=classifier_lists.headers.index('Decision:') + 1, row=new_row).value = decision  # Change decision and school fields
			ws.cell(column=classifier_lists.headers.index('College:') + 1, row=new_row).value = school
			counter += 1
	print('Additional Entries done')
